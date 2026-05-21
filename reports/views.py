from io import BytesIO

from django.db import models
from django.db.models import Count, Sum
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone
from django.views.decorators.clickjacking import xframe_options_sameorigin
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from backoffice.decorators import staff_required
from backoffice.views import _require_any_permission
from catalog.models import StockItem
from orders.models import Order
from .documents import (
    DOCUMENT_LABELS,
    build_document_pdf_bytes,
    build_documents_zip,
    document_filename,
)


@staff_required
def dashboard(request):
    _require_any_permission(request.user, 'orders.view_order')
    orders = Order.objects.prefetch_related('items')
    revenue = orders.filter(payment_status=Order.PAYMENT_PAID).exclude(status=Order.STATUS_CANCELLED).aggregate(
        total=Sum('subtotal')
    )['total'] or 0
    status_labels = dict(Order.STATUS_CHOICES)
    payment_labels = dict(Order.PAYMENT_STATUS_CHOICES)
    status_rows = [
        {**row, 'label': status_labels.get(row['status'], row['status'])}
        for row in orders.values('status').annotate(count=Count('id'), total=Sum('subtotal')).order_by('status')
    ]
    payment_rows = [
        {**row, 'label': payment_labels.get(row['payment_status'], row['payment_status'])}
        for row in orders.values('payment_status').annotate(count=Count('id'), total=Sum('subtotal')).order_by('payment_status')
    ]
    context = {
        'orders_total': orders.count(),
        'orders_new': orders.filter(status=Order.STATUS_NEW).count(),
        'orders_confirmed': orders.filter(status=Order.STATUS_CONFIRMED).count(),
        'orders_completed': orders.filter(status=Order.STATUS_COMPLETED).count(),
        'payments_paid': orders.filter(payment_status=Order.PAYMENT_PAID).count(),
        'payments_waiting': orders.filter(payment_status=Order.PAYMENT_WAITING).count(),
        'payments_failed': orders.filter(payment_status=Order.PAYMENT_FAILED).count(),
        'revenue': revenue,
        'low_stock': StockItem.objects.select_related('product').filter(
            quantity__lte=models.F('min_quantity') + models.F('reserved_quantity')
        )[:20],
        'status_rows': status_rows,
        'payment_rows': payment_rows,
    }
    return render(request, 'reports/dashboard.html', context)


def _xlsx_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _style_header(sheet):
    fill = PatternFill('solid', fgColor='1F2937')
    font = Font(color='FFFFFF', bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font


@staff_required
def sales_report_xlsx(request):
    _require_any_permission(request.user, 'orders.view_order')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Продажи'
    sheet.append(['ID', 'Дата', 'Статус', 'Оплата', 'Способ оплаты', 'Дата оплаты', 'Платеж', 'Клиент', 'Телефон', 'Сумма'])
    for order in Order.objects.order_by('-created_at'):
        sheet.append(
            [
                order.pk,
                timezone.localtime(order.created_at).strftime('%d.%m.%Y %H:%M'),
                order.get_status_display(),
                order.get_payment_status_display(),
                order.get_payment_method_display(),
                timezone.localtime(order.paid_at).strftime('%d.%m.%Y %H:%M') if order.paid_at else '',
                order.payment_reference,
                order.customer_name,
                order.customer_phone,
                float(order.subtotal),
            ]
        )
    _style_header(sheet)
    return _xlsx_response(workbook, 'doorsky-sales-report.xlsx')


@staff_required
def stock_report_xlsx(request):
    _require_any_permission(request.user, 'catalog.view_stockitem')
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Склад'
    sheet.append(['Артикул', 'Товар', 'Категория', 'На складе', 'В резерве', 'Доступно', 'Мин. остаток'])
    for stock in StockItem.objects.select_related('product', 'product__category').order_by('product__sku'):
        sheet.append(
            [
                stock.product.sku,
                stock.product.name,
                stock.product.category.name,
                stock.quantity,
                stock.reserved_quantity,
                stock.available_quantity,
                stock.min_quantity,
            ]
        )
    _style_header(sheet)
    return _xlsx_response(workbook, 'doorsky-stock-report.xlsx')


def _has_document_access(request, order, public_key=None):
    has_public_key = public_key and order.public_key == public_key
    is_owner = request.user.is_authenticated and order.user_id == request.user.id
    has_staff_access = request.user.is_staff and request.user.has_perm('orders.can_export_order_docs')
    return has_staff_access or is_owner or has_public_key


def _file_response(content, filename, content_type, disposition='attachment'):
    response = HttpResponse(
        content,
        content_type=content_type,
    )
    response['Content-Disposition'] = f'{disposition}; filename="{filename}"'
    return response


@xframe_options_sameorigin
def order_document_pdf(request, pk, document_type, public_key=None):
    order = get_object_or_404(Order.objects.prefetch_related('items__product'), pk=pk)
    if not _has_document_access(request, order, public_key):
        raise Http404

    if document_type not in DOCUMENT_LABELS:
        raise Http404
    return _file_response(
        build_document_pdf_bytes(order, document_type),
        document_filename(order, document_type, 'pdf'),
        'application/pdf',
        'attachment' if request.GET.get('download') == '1' else 'inline',
    )


def order_documents_zip(request, pk, public_key=None):
    order = get_object_or_404(Order.objects.prefetch_related('items__product'), pk=pk)
    if not _has_document_access(request, order, public_key):
        raise Http404

    response = HttpResponse(
        build_documents_zip(order),
        content_type='application/zip',
    )
    response['Content-Disposition'] = f'attachment; filename="doorsky-pdf-documents-{order.pk}.zip"'
    return response
